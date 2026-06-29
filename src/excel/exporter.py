"""
Holonet

Excel Exporter
"""

from __future__ import annotations

import pandas as pd

from src.database.connection import DatabaseConnection
from src.utils.logger import get_logger
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill

logger = get_logger(__name__)


class ExcelExporter:

    def load_report(self) -> pd.DataFrame:
        """
        Loads the report from SQL Server.
        """
        logger.info(
            "Loading report data..."
        )

        with DatabaseConnection() as connection:

            #
            # Load SQL Query
            #

            sql_file = (
                Path(__file__).parent.parent
                / "sql"
                / "service_line_summary_v3.sql"
            )

            logger.info(
                f"Loading SQL query: {sql_file}"
            )

            report_query = sql_file.read_text(
                encoding="utf-8"
            )

            #
            # Execute Query
            #

            dataframe = pd.read_sql(
                report_query,
                connection
            )
            dataframe.rename(
                columns={
                    "account_number": "Account Number",
                    "service_line_number": "Service Line",
                    "nickname": "Nickname",

                    "product_name": "Plan",
                    "product_id": "Product ID",

                    "billing_cycle_key": "Billing Cycle",
                    "cycle_start": "Cycle Start",
                    "cycle_end": "Cycle End",

                    "contracted_capacity_gb": "Contracted Capacity (GB)",
                    "topup_capacity_gb": "TopUp Capacity (GB)",
                    "total_capacity_gb": "Total Capacity (GB)",

                    "total_consumed_gb": "Consumed (GB)",
                    "available_gb": "Available (GB)",

                    "current_usage_percent": "Current Usage (%)",
                    "contract_usage_percent": "Contract Usage (%)",

                    "topup_quantity": "TopUp Quantity",
                    "topup_consumed_gb": "TopUp Used (GB)",
                    "topup_remaining_gb": "TopUp Remaining (GB)",
                    "topup_cost": "TopUp Cost",

                    "topup_required": "TopUp Required",

                    "recurring_cost": "Monthly Cost",

                    "currency": "Currency",

                    "current_status": "Current Status",
                    "contract_status": "Contract Status",

                    "last_updated": "Last Updated"
                },
                inplace=True
            )

        logger.info(
            f"{len(dataframe)} rows loaded."
        )

        return dataframe
    
    def export_report(self) -> Path:
        """
        Exports the report to Excel.
        """
        dataframe = self.load_report()

        exports_folder = Path("exports")
        exports_folder.mkdir(
            exist_ok=True
        )

        file_name = (
            f"Holonet_"
            f"{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        output_file = exports_folder / file_name

        logger.info(
            f"Generating Excel: {output_file}"
        )

        dataframe.to_excel(
            output_file,
            index=False
        )

        workbook = load_workbook(output_file)
        worksheet = workbook.active

        header_fill = PatternFill(
            fill_type="solid",
            start_color="C32032",
            end_color="C32032"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        for column_cells in worksheet.columns:
            length = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            adjusted_width = min(length + 3, 60)
            worksheet.column_dimensions[
                column_cells[0].column_letter
            ].width = adjusted_width

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        gb_columns = [
            "Contracted Capacity (GB)",
            "TopUp Capacity (GB)",
            "Total Capacity (GB)",
            "Consumed (GB)",
            "Available (GB)",
            "TopUp Used (GB)",
            "TopUp Remaining (GB)"
        ]

        currency_columns = [
            "Monthly Cost",
            "TopUp Cost"
        ]

        date_columns = [
            "Cycle Start",
            "Cycle End",
            "Last Updated"
        ]

        percent_columns = [
            "Current Usage (%)",
            "Contract Usage (%)"
        ]

        header_map = {}
        for cell in worksheet[1]:
            header_map[cell.value] = cell.column

        for column_name in gb_columns:
            if column_name in header_map:
                column = header_map[column_name]
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=row,
                        column=column
                    ).number_format = "0.00"

        for column_name in currency_columns:
            if column_name in header_map:
                column = header_map[column_name]
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=row,
                        column=column
                    ).number_format = "$#,##0.00"

        for column_name in date_columns:
            if column_name in header_map:
                column = header_map[column_name]
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=row,
                        column=column
                    ).number_format = "yyyy-mm-dd"

        for column_name in percent_columns:
            if column_name in header_map:
                column = header_map[column_name]
                for row in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=row,
                        column=column
                    ).number_format = '0.00"%"'

        # ----------------------------------------------------------------------
        # Color Contract Usage % (Corregido: fuera del bucle de percent_columns)
        # ----------------------------------------------------------------------
        if "Contract Usage (%)" in header_map:
            column = header_map["Contract Usage (%)"]
            for row in range(2, worksheet.max_row + 1):
                cell = worksheet.cell(
                    row=row,
                    column=column
                )

                if cell.value is None:
                    continue

                if cell.value >= 100:
                    cell.font = Font(
                        bold=True,
                        color="C00000"
                    )
                elif cell.value >= 80:
                    cell.font = Font(
                        bold=True,
                        color="C09000"
                    )
                else:
                    cell.font = Font(
                        bold=True,
                        color="008000"
                    )

        workbook.save(output_file)

        logger.info(
            "Excel generated successfully."
        )

        return output_file