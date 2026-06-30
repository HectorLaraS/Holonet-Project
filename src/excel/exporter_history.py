"""
Holonet

History Excel Exporter (Version 1)
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from src.utils.logger import get_logger

logger = get_logger(__name__)


class HistoryExporter:

    def __init__(self, usage_data: dict):
        self.usage_data = usage_data

    def build_dataframe(self) -> pd.DataFrame:
        rows = []

        for service in self.usage_data["content"]["results"]:

            row = {
                "Account Number": service.get("accountNumber"),
                "Service Line": service.get("serviceLineNumber"),
                "Nickname": service.get("nickname", ""),
                "Plan": service.get("productReferenceId", ""),
            }

            billing_cycles = sorted(
                service.get("billingCycles", []),
                key=lambda x: x["startDate"],
                reverse=True
            )

            labels = ["Current Month", "Last Month", "Month -2", "Month -3"]

            for label, cycle in zip(labels, billing_cycles):

                contracted = 0.0
                topup_qty = 0
                topup_capacity = 0.0
                topup_used = 0.0
                topup_cost = 0.0

                for pool in cycle.get("dataPoolUsage", []):
                    for block in pool.get("dataBlocks", []):
                        t = block.get("dataBlockType")

                        if t == "RecurringPerBillingCycle":
                            contracted += block.get("totalAmountGB", 0)

                        elif t == "Overage":
                            topup_qty += block.get("blocksCount", 0)
                            topup_capacity += block.get("totalAmountGB", 0)
                            topup_used += block.get("consumedAmountGB", 0)
                            topup_cost += block.get("totalPrice", 0)

                consumed = cycle.get("totalPriorityGB", 0)

                usage = round(consumed * 100 / contracted, 2) if contracted else 0

                row[f"{label} Contracted (GB)"] = contracted
                row[f"{label} Used (GB)"] = consumed
                row[f"{label} Usage (%)"] = usage
                row[f"{label} TopUps"] = topup_qty
                row[f"{label} TopUp Capacity (GB)"] = topup_capacity
                row[f"{label} TopUp Used (GB)"] = topup_used
                row[f"{label} TopUp Remaining (GB)"] = topup_capacity - topup_used
                row[f"{label} TopUp Cost"] = topup_cost

            rows.append(row)

        return pd.DataFrame(rows)
    
    def build_chart_dataframe(self) -> pd.DataFrame:
        """
        Builds a normalized dataset optimized for
        Power BI and Excel Pivot Charts.
        """

        rows = []

        periods = [
            ("Current Month", 0),
            ("Last Month", 1),
            ("Month -2", 2),
            ("Month -3", 3)
        ]

        for service in self.usage_data["content"]["results"]:

            billing_cycles = sorted(
                service.get("billingCycles", []),
                key=lambda x: x["startDate"],
                reverse=True
            )

            base = {
                "Account Number": service.get("accountNumber"),
                "Service Line": service.get("serviceLineNumber"),
                "Nickname": service.get("nickname", ""),
                "Plan": service.get("productReferenceId", "")
            }

            for (period_name, period_order), cycle in zip(
                periods,
                billing_cycles
            ):

                contracted = 0.0
                topup_qty = 0
                topup_capacity = 0.0
                topup_used = 0.0
                topup_cost = 0.0

                for pool in cycle.get("dataPoolUsage", []):

                    for block in pool.get("dataBlocks", []):

                        block_type = block.get("dataBlockType")

                        if block_type == "RecurringPerBillingCycle":

                            contracted += block.get(
                                "totalAmountGB",
                                0
                            )

                        elif block_type == "Overage":

                            topup_qty += block.get(
                                "blocksCount",
                                0
                            )

                            topup_capacity += block.get(
                                "totalAmountGB",
                                0
                            )

                            topup_used += block.get(
                                "consumedAmountGB",
                                0
                            )

                            topup_cost += block.get(
                                "totalPrice",
                                0
                            )

                consumed = cycle.get(
                    "totalPriorityGB",
                    0
                )

                usage = (
                    round(
                        consumed * 100 / contracted,
                        2
                    )
                    if contracted
                    else 0
                )

                metrics = {
                    "Contracted (GB)": contracted,
                    "Used (GB)": consumed,
                    "Usage (%)": usage,
                    "TopUps": topup_qty,
                    "TopUp Capacity (GB)": topup_capacity,
                    "TopUp Used (GB)": topup_used,
                    "TopUp Remaining (GB)": (
                        topup_capacity - topup_used
                    ),
                    "TopUp Cost": topup_cost
                }

                for metric, value in metrics.items():

                    row = base.copy()

                    row["Period"] = period_name
                    row["Period Order"] = period_order
                    row["Metric"] = metric
                    row["Value"] = value

                    rows.append(row)

        dataframe = pd.DataFrame(rows)

        return dataframe

    def export_report(self) -> Path:

        summary_df = self.build_dataframe()

        chart_df = self.build_chart_dataframe()

        exports_folder = Path("exports") / "history"
        exports_folder.mkdir(parents=True, exist_ok=True)

        output_file = exports_folder / (
            f"Holonet_History_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
        )

        with pd.ExcelWriter(
            output_file,
            engine="openpyxl"
        ) as writer:

            summary_df.to_excel(
                writer,
                sheet_name="Summary",
                index=False
            )

            chart_df.to_excel(
                writer,
                sheet_name="Chart_Data",
                index=False
            )

        workbook = load_workbook(output_file)

        fill = PatternFill(fill_type="solid", start_color="C32032", end_color="C32032")
        font = Font(bold=True, color="FFFFFF")
        align = Alignment(horizontal="center", vertical="center")

        for worksheet in workbook.worksheets:

            for cell in worksheet[1]:
                cell.fill = fill
                cell.font = font
                cell.alignment = align

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions

        return output_file
